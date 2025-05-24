"""
Система экспорта данных для ППСД
"""

from PySide6.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QLabel, 
                             QPushButton, QComboBox, QCheckBox, QGroupBox,
                             QGridLayout, QProgressBar, QTextEdit, QFileDialog,
                             QDateEdit, QSpinBox, QFrame, QTabWidget, QWidget,
                             QListWidget, QListWidgetItem, QMessageBox)
from PySide6.QtCore import Qt, QThread, Signal, QDate, QTimer
from PySide6.QtGui import QFont, QPixmap
from datetime import datetime, timedelta
import os
import json
from ui.themes import theme_manager
from ui.icons.icon_provider import IconProvider

class ExportWorker(QThread):
    """Рабочий поток для экспорта данных"""
    
    progress_updated = Signal(int)
    status_updated = Signal(str)
    export_completed = Signal(str)  # Путь к файлу
    export_failed = Signal(str)     # Сообщение об ошибке
    
    def __init__(self, export_config):
        super().__init__()
        self.export_config = export_config
        self.is_cancelled = False
    
    def cancel(self):
        """Отмена экспорта"""
        self.is_cancelled = True
    
    def run(self):
        """Выполнение экспорта"""
        try:
            self.status_updated.emit("Подготовка данных...")
            self.progress_updated.emit(10)
            
            if self.is_cancelled:
                return
            
            # Здесь будет логика экспорта в зависимости от формата
            format_type = self.export_config.get('format', 'excel')
            
            if format_type == 'excel':
                self.export_to_excel()
            elif format_type == 'pdf':
                self.export_to_pdf()
            elif format_type == 'csv':
                self.export_to_csv()
            elif format_type == 'json':
                self.export_to_json()
            
        except Exception as e:
            self.export_failed.emit(str(e))
    
    def export_to_excel(self):
        """Экспорт в Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            self.status_updated.emit("Создание Excel файла...")
            self.progress_updated.emit(30)
            
            # Создаем рабочую книгу
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Материалы"
            
            # Заголовки
            headers = ["ID", "Марка материала", "Тип", "Партия", "Плавка", 
                      "Поставщик", "Статус", "Дата создания"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            self.progress_updated.emit(50)
            
            # Данные (здесь должна быть логика получения данных из БД)
            # Для примера добавим несколько строк
            sample_data = [
                [1, "08Х18Н10Т", "Круг", "Б001", "П001", "ООО Поставщик", "На складе", "2024-01-15"],
                [2, "12Х18Н10Т", "Лист", "Б002", "П002", "ООО Поставщик 2", "В производстве", "2024-01-16"]
            ]
            
            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                
                if self.is_cancelled:
                    return
            
            self.progress_updated.emit(80)
            
            # Автоширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Сохранение
            file_path = self.export_config.get('file_path', 'export.xlsx')
            wb.save(file_path)
            
            self.progress_updated.emit(100)
            self.status_updated.emit("Экспорт завершен")
            self.export_completed.emit(file_path)
            
        except ImportError:
            self.export_failed.emit("Модуль openpyxl не установлен. Установите: pip install openpyxl")
        except Exception as e:
            self.export_failed.emit(f"Ошибка при экспорте в Excel: {str(e)}")
    
    def export_to_pdf(self):
        """Экспорт в PDF"""
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            self.status_updated.emit("Создание PDF файла...")
            self.progress_updated.emit(30)
            
            file_path = self.export_config.get('file_path', 'export.pdf')
            doc = SimpleDocTemplate(file_path, pagesize=landscape(A4))
            
            # Стили
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Центрирование
            )
            
            # Элементы документа
            elements = []
            
            # Заголовок
            title = Paragraph("Отчет по материалам ППСД", title_style)
            elements.append(title)
            elements.append(Spacer(1, 20))
            
            self.progress_updated.emit(50)
            
            # Таблица данных
            data = [
                ["ID", "Марка", "Тип", "Партия", "Плавка", "Поставщик", "Статус", "Дата"],
                ["1", "08Х18Н10Т", "Круг", "Б001", "П001", "ООО Поставщик", "На складе", "15.01.2024"],
                ["2", "12Х18Н10Т", "Лист", "Б002", "П002", "ООО Поставщик 2", "В производстве", "16.01.2024"]
            ]
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
            
            self.progress_updated.emit(80)
            
            # Построение документа
            doc.build(elements)
            
            self.progress_updated.emit(100)
            self.status_updated.emit("Экспорт завершен")
            self.export_completed.emit(file_path)
            
        except ImportError:
            self.export_failed.emit("Модуль reportlab не установлен. Установите: pip install reportlab")
        except Exception as e:
            self.export_failed.emit(f"Ошибка при экспорте в PDF: {str(e)}")
    
    def export_to_csv(self):
        """Экспорт в CSV"""
        try:
            import csv
            
            self.status_updated.emit("Создание CSV файла...")
            self.progress_updated.emit(30)
            
            file_path = self.export_config.get('file_path', 'export.csv')
            
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile, delimiter=';')
                
                # Заголовки
                headers = ["ID", "Марка материала", "Тип", "Партия", "Плавка", 
                          "Поставщик", "Статус", "Дата создания"]
                writer.writerow(headers)
                
                self.progress_updated.emit(50)
                
                # Данные
                sample_data = [
                    [1, "08Х18Н10Т", "Круг", "Б001", "П001", "ООО Поставщик", "На складе", "2024-01-15"],
                    [2, "12Х18Н10Т", "Лист", "Б002", "П002", "ООО Поставщик 2", "В производстве", "2024-01-16"]
                ]
                
                for row in sample_data:
                    writer.writerow(row)
                    if self.is_cancelled:
                        return
                
                self.progress_updated.emit(100)
            
            self.status_updated.emit("Экспорт завершен")
            self.export_completed.emit(file_path)
            
        except Exception as e:
            self.export_failed.emit(f"Ошибка при экспорте в CSV: {str(e)}")
    
    def export_to_json(self):
        """Экспорт в JSON"""
        try:
            self.status_updated.emit("Создание JSON файла...")
            self.progress_updated.emit(30)
            
            # Подготовка данных
            data = {
                "export_info": {
                    "timestamp": datetime.now().isoformat(),
                    "format": "json",
                    "version": "1.0"
                },
                "materials": [
                    {
                        "id": 1,
                        "material_grade": "08Х18Н10Т",
                        "type": "Круг",
                        "batch": "Б001",
                        "melt": "П001",
                        "supplier": "ООО Поставщик",
                        "status": "На складе",
                        "created_at": "2024-01-15"
                    },
                    {
                        "id": 2,
                        "material_grade": "12Х18Н10Т",
                        "type": "Лист",
                        "batch": "Б002",
                        "melt": "П002",
                        "supplier": "ООО Поставщик 2",
                        "status": "В производстве",
                        "created_at": "2024-01-16"
                    }
                ]
            }
            
            self.progress_updated.emit(70)
            
            file_path = self.export_config.get('file_path', 'export.json')
            
            with open(file_path, 'w', encoding='utf-8') as jsonfile:
                json.dump(data, jsonfile, ensure_ascii=False, indent=2)
            
            self.progress_updated.emit(100)
            self.status_updated.emit("Экспорт завершен")
            self.export_completed.emit(file_path)
            
        except Exception as e:
            self.export_failed.emit(f"Ошибка при экспорте в JSON: {str(e)}")

class ExportDialog(QDialog):
    """Диалог экспорта данных"""
    
    def __init__(self, parent=None, data_type="materials"):
        super().__init__(parent)
        self.data_type = data_type
        self.export_worker = None
        self.init_ui()
        self.apply_styles()
    
    def init_ui(self):
        """Инициализация интерфейса"""
        self.setWindowTitle("Экспорт данных")
        self.setFixedSize(600, 500)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(16)
        
        # Заголовок
        title_label = QLabel("Экспорт данных")
        title_label.setFont(QFont("Segoe UI", 16, QFont.Weight.Bold))
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        # Вкладки
        tab_widget = QTabWidget()
        layout.addWidget(tab_widget)
        
        # Вкладка "Формат"
        format_tab = self.create_format_tab()
        tab_widget.addTab(format_tab, "Формат")
        
        # Вкладка "Фильтры"
        filters_tab = self.create_filters_tab()
        tab_widget.addTab(filters_tab, "Фильтры")
        
        # Вкладка "Настройки"
        settings_tab = self.create_settings_tab()
        tab_widget.addTab(settings_tab, "Настройки")
        
        # Прогресс
        progress_group = QGroupBox("Прогресс экспорта")
        progress_layout = QVBoxLayout(progress_group)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        progress_layout.addWidget(self.progress_bar)
        
        self.status_label = QLabel("Готов к экспорту")
        self.status_label.setStyleSheet(f"color: {theme_manager.get_color('text_secondary')};")
        progress_layout.addWidget(self.status_label)
        
        layout.addWidget(progress_group)
        
        # Кнопки
        buttons_layout = QHBoxLayout()
        
        self.preview_btn = QPushButton("Предпросмотр")
        self.preview_btn.setIcon(IconProvider.create_search_icon())
        self.preview_btn.clicked.connect(self.show_preview)
        self.preview_btn.setStyleSheet(theme_manager.get_button_style('neutral'))
        buttons_layout.addWidget(self.preview_btn)
        
        buttons_layout.addStretch()
        
        self.cancel_btn = QPushButton("Отмена")
        self.cancel_btn.clicked.connect(self.cancel_export)
        self.cancel_btn.setStyleSheet(theme_manager.get_button_style('neutral'))
        buttons_layout.addWidget(self.cancel_btn)
        
        self.export_btn = QPushButton("Экспорт")
        self.export_btn.setIcon(IconProvider.create_excel_icon())
        self.export_btn.clicked.connect(self.start_export)
        self.export_btn.setStyleSheet(theme_manager.get_button_style('primary'))
        buttons_layout.addWidget(self.export_btn)
        
        layout.addLayout(buttons_layout)
    
    def create_format_tab(self):
        """Создание вкладки выбора формата"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Формат файла
        format_group = QGroupBox("Формат файла")
        format_layout = QGridLayout(format_group)
        
        self.format_combo = QComboBox()
        self.format_combo.addItems([
            "Excel (.xlsx)",
            "PDF (.pdf)", 
            "CSV (.csv)",
            "JSON (.json)"
        ])
        self.format_combo.currentTextChanged.connect(self.on_format_changed)
        format_layout.addWidget(QLabel("Формат:"), 0, 0)
        format_layout.addWidget(self.format_combo, 0, 1)
        
        # Путь к файлу
        self.file_path_edit = QTextEdit()
        self.file_path_edit.setPlaceholderText("Выберите путь для сохранения...")
        format_layout.addWidget(QLabel("Путь:"), 1, 0)
        format_layout.addWidget(self.file_path_edit, 1, 1)
        
        browse_btn = QPushButton("Обзор...")
        browse_btn.clicked.connect(self.browse_file_path)
        browse_btn.setStyleSheet(theme_manager.get_button_style('neutral'))
        format_layout.addWidget(browse_btn, 1, 2)
        
        layout.addWidget(format_group)
        
        # Опции формата
        self.format_options_group = QGroupBox("Опции формата")
        self.format_options_layout = QVBoxLayout(self.format_options_group)
        layout.addWidget(self.format_options_group)
        
        self.update_format_options()
        
        layout.addStretch()
        return widget
    
    def create_filters_tab(self):
        """Создание вкладки фильтров"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Период
        period_group = QGroupBox("Период")
        period_layout = QGridLayout(period_group)
        
        self.date_from = QDateEdit()
        self.date_from.setDate(QDate.currentDate().addDays(-30))
        self.date_from.setCalendarPopup(True)
        period_layout.addWidget(QLabel("С:"), 0, 0)
        period_layout.addWidget(self.date_from, 0, 1)
        
        self.date_to = QDateEdit()
        self.date_to.setDate(QDate.currentDate())
        self.date_to.setCalendarPopup(True)
        period_layout.addWidget(QLabel("По:"), 0, 2)
        period_layout.addWidget(self.date_to, 0, 3)
        
        layout.addWidget(period_group)
        
        # Статусы
        status_group = QGroupBox("Статусы материалов")
        status_layout = QVBoxLayout(status_group)
        
        self.status_checkboxes = {}
        statuses = ["На складе", "В производстве", "В испытаниях", "Завершен", "Отклонен"]
        
        for status in statuses:
            checkbox = QCheckBox(status)
            checkbox.setChecked(True)
            self.status_checkboxes[status] = checkbox
            status_layout.addWidget(checkbox)
        
        layout.addWidget(status_group)
        
        # Дополнительные фильтры
        additional_group = QGroupBox("Дополнительные фильтры")
        additional_layout = QVBoxLayout(additional_group)
        
        self.include_deleted = QCheckBox("Включить удаленные записи")
        additional_layout.addWidget(self.include_deleted)
        
        self.only_with_samples = QCheckBox("Только с образцами")
        additional_layout.addWidget(self.only_with_samples)
        
        layout.addWidget(additional_group)
        
        layout.addStretch()
        return widget
    
    def create_settings_tab(self):
        """Создание вкладки настроек"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Колонки для экспорта
        columns_group = QGroupBox("Колонки для экспорта")
        columns_layout = QVBoxLayout(columns_group)
        
        self.column_checkboxes = {}
        columns = [
            ("ID", "id"),
            ("Марка материала", "material_grade"),
            ("Тип продукта", "product_type"),
            ("Номер партии", "batch_number"),
            ("Номер плавки", "melt_number"),
            ("Поставщик", "supplier"),
            ("Статус", "status"),
            ("Дата создания", "created_at"),
            ("Дата обновления", "updated_at")
        ]
        
        for display_name, field_name in columns:
            checkbox = QCheckBox(display_name)
            checkbox.setChecked(True)
            self.column_checkboxes[field_name] = checkbox
            columns_layout.addWidget(checkbox)
        
        layout.addWidget(columns_group)
        
        # Настройки экспорта
        export_settings_group = QGroupBox("Настройки экспорта")
        export_settings_layout = QGridLayout(export_settings_group)
        
        # Лимит записей
        export_settings_layout.addWidget(QLabel("Лимит записей:"), 0, 0)
        self.limit_spinbox = QSpinBox()
        self.limit_spinbox.setRange(1, 100000)
        self.limit_spinbox.setValue(10000)
        export_settings_layout.addWidget(self.limit_spinbox, 0, 1)
        
        # Кодировка
        export_settings_layout.addWidget(QLabel("Кодировка:"), 1, 0)
        self.encoding_combo = QComboBox()
        self.encoding_combo.addItems(["UTF-8", "Windows-1251", "CP866"])
        export_settings_layout.addWidget(self.encoding_combo, 1, 1)
        
        layout.addWidget(export_settings_group)
        
        layout.addStretch()
        return widget
    
    def on_format_changed(self):
        """Обработка изменения формата"""
        self.update_format_options()
        self.update_file_extension()
    
    def update_format_options(self):
        """Обновление опций формата"""
        # Очищаем предыдущие опции
        for i in reversed(range(self.format_options_layout.count())):
            self.format_options_layout.itemAt(i).widget().setParent(None)
        
        format_text = self.format_combo.currentText()
        
        if "Excel" in format_text:
            # Опции для Excel
            self.include_charts = QCheckBox("Включить диаграммы")
            self.format_options_layout.addWidget(self.include_charts)
            
            self.auto_width = QCheckBox("Автоширина колонок")
            self.auto_width.setChecked(True)
            self.format_options_layout.addWidget(self.auto_width)
            
        elif "PDF" in format_text:
            # Опции для PDF
            self.landscape_mode = QCheckBox("Альбомная ориентация")
            self.landscape_mode.setChecked(True)
            self.format_options_layout.addWidget(self.landscape_mode)
            
            self.include_logo = QCheckBox("Включить логотип")
            self.format_options_layout.addWidget(self.include_logo)
            
        elif "CSV" in format_text:
            # Опции для CSV
            delimiter_layout = QHBoxLayout()
            delimiter_layout.addWidget(QLabel("Разделитель:"))
            self.delimiter_combo = QComboBox()
            self.delimiter_combo.addItems([";", ",", "\t"])
            delimiter_layout.addWidget(self.delimiter_combo)
            delimiter_layout.addStretch()
            self.format_options_layout.addLayout(delimiter_layout)
    
    def update_file_extension(self):
        """Обновление расширения файла"""
        current_path = self.file_path_edit.text()
        if current_path:
            base_path = os.path.splitext(current_path)[0]
        else:
            base_path = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        format_text = self.format_combo.currentText()
        
        if "Excel" in format_text:
            extension = ".xlsx"
        elif "PDF" in format_text:
            extension = ".pdf"
        elif "CSV" in format_text:
            extension = ".csv"
        elif "JSON" in format_text:
            extension = ".json"
        else:
            extension = ".txt"
        
        self.file_path_edit.setText(base_path + extension)
    
    def browse_file_path(self):
        """Выбор пути для сохранения файла"""
        format_text = self.format_combo.currentText()
        
        if "Excel" in format_text:
            filter_str = "Excel files (*.xlsx);;All files (*.*)"
            default_name = "export.xlsx"
        elif "PDF" in format_text:
            filter_str = "PDF files (*.pdf);;All files (*.*)"
            default_name = "export.pdf"
        elif "CSV" in format_text:
            filter_str = "CSV files (*.csv);;All files (*.*)"
            default_name = "export.csv"
        elif "JSON" in format_text:
            filter_str = "JSON files (*.json);;All files (*.*)"
            default_name = "export.json"
        else:
            filter_str = "All files (*.*)"
            default_name = "export.txt"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить как", default_name, filter_str
        )
        
        if file_path:
            self.file_path_edit.setText(file_path)
    
    def show_preview(self):
        """Показать предпросмотр данных"""
        preview_dialog = QDialog(self)
        preview_dialog.setWindowTitle("Предпросмотр данных")
        preview_dialog.setFixedSize(800, 600)
        
        layout = QVBoxLayout(preview_dialog)
        
        preview_text = QTextEdit()
        preview_text.setReadOnly(True)
        
        # Пример данных для предпросмотра
        preview_content = """
Предпросмотр экспорта данных:

Формат: {format}
Период: {date_from} - {date_to}
Количество записей: ~{count}

Пример данных:
ID | Марка материала | Тип | Партия | Плавка | Поставщик | Статус | Дата
1  | 08Х18Н10Т     | Круг | Б001   | П001   | ООО Поставщик | На складе | 15.01.2024
2  | 12Х18Н10Т     | Лист | Б002   | П002   | ООО Поставщик 2 | В производстве | 16.01.2024
...
        """.format(
            format=self.format_combo.currentText(),
            date_from=self.date_from.date().toString("dd.MM.yyyy"),
            date_to=self.date_to.date().toString("dd.MM.yyyy"),
            count=self.limit_spinbox.value()
        )
        
        preview_text.setPlainText(preview_content)
        layout.addWidget(preview_text)
        
        close_btn = QPushButton("Закрыть")
        close_btn.clicked.connect(preview_dialog.close)
        close_btn.setStyleSheet(theme_manager.get_button_style('neutral'))
        layout.addWidget(close_btn)
        
        preview_dialog.exec()
    
    def start_export(self):
        """Начать экспорт"""
        if not self.file_path_edit.text():
            QMessageBox.warning(self, "Предупреждение", "Выберите путь для сохранения файла")
            return
        
        # Подготовка конфигурации экспорта
        export_config = {
            'format': self.get_format_key(),
            'file_path': self.file_path_edit.text(),
            'date_from': self.date_from.date().toPython(),
            'date_to': self.date_to.date().toPython(),
            'limit': self.limit_spinbox.value(),
            'encoding': self.encoding_combo.currentText(),
            'columns': [field for field, checkbox in self.column_checkboxes.items() if checkbox.isChecked()],
            'statuses': [status for status, checkbox in self.status_checkboxes.items() if checkbox.isChecked()]
        }
        
        # Запуск экспорта в отдельном потоке
        self.export_worker = ExportWorker(export_config)
        self.export_worker.progress_updated.connect(self.progress_bar.setValue)
        self.export_worker.status_updated.connect(self.status_label.setText)
        self.export_worker.export_completed.connect(self.on_export_completed)
        self.export_worker.export_failed.connect(self.on_export_failed)
        
        # Обновление интерфейса
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.export_btn.setEnabled(False)
        self.cancel_btn.setText("Отменить экспорт")
        
        self.export_worker.start()
    
    def get_format_key(self):
        """Получить ключ формата"""
        format_text = self.format_combo.currentText()
        if "Excel" in format_text:
            return "excel"
        elif "PDF" in format_text:
            return "pdf"
        elif "CSV" in format_text:
            return "csv"
        elif "JSON" in format_text:
            return "json"
        return "excel"
    
    def cancel_export(self):
        """Отмена экспорта"""
        if self.export_worker and self.export_worker.isRunning():
            self.export_worker.cancel()
            self.export_worker.wait()
            self.status_label.setText("Экспорт отменен")
            self.reset_ui()
        else:
            self.reject()
    
    def on_export_completed(self, file_path):
        """Обработка завершения экспорта"""
        self.reset_ui()
        
        msg = QMessageBox(self)
        msg.setWindowTitle("Экспорт завершен")
        msg.setText(f"Данные успешно экспортированы в файл:\n{file_path}")
        msg.setIcon(QMessageBox.Icon.Information)
        
        open_btn = msg.addButton("Открыть файл", QMessageBox.ButtonRole.ActionRole)
        open_folder_btn = msg.addButton("Открыть папку", QMessageBox.ButtonRole.ActionRole)
        close_btn = msg.addButton("Закрыть", QMessageBox.ButtonRole.RejectRole)
        
        msg.exec()
        
        if msg.clickedButton() == open_btn:
            os.startfile(file_path)
        elif msg.clickedButton() == open_folder_btn:
            os.startfile(os.path.dirname(file_path))
    
    def on_export_failed(self, error_message):
        """Обработка ошибки экспорта"""
        self.reset_ui()
        QMessageBox.critical(self, "Ошибка экспорта", error_message)
    
    def reset_ui(self):
        """Сброс интерфейса после экспорта"""
        self.progress_bar.setVisible(False)
        self.export_btn.setEnabled(True)
        self.cancel_btn.setText("Отмена")
    
    def apply_styles(self):
        """Применение стилей"""
        colors = theme_manager.get_current_theme()['colors']
        
        self.setStyleSheet(f"""
        QDialog {{
            background-color: {colors['background']};
            color: {colors['text_primary']};
        }}
        QGroupBox {{
            font-weight: bold;
            border: 1px solid {colors['border']};
            border-radius: 6px;
            margin-top: 8px;
            padding-top: 8px;
        }}
        QGroupBox::title {{
            subcontrol-origin: margin;
            left: 10px;
            padding: 0 5px 0 5px;
        }}
        QTabWidget::pane {{
            border: 1px solid {colors['border']};
            border-radius: 4px;
        }}
        QTabBar::tab {{
            background-color: {colors['card']};
            border: 1px solid {colors['border']};
            padding: 8px 16px;
            margin-right: 2px;
        }}
        QTabBar::tab:selected {{
            background-color: {colors['hover']};
            border-bottom-color: {colors['hover']};
        }}
        """) 